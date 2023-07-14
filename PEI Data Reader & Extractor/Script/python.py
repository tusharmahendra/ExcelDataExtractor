#!/usr/bin/env python
# coding: utf-8

# <h2>Hello!</h2>
# 
# <p>I am a PEI Finance Bot. I am designed to manipulate big data of all stores. I can uniquly identify data for each store. I use my super ability to write each store's PEI data into their excel workbook. 
#     
# Don't worry if you're a new store and your file doesn't exist or you deleted your file by mistake! I will automatically generate one for you.
#     
# <i>author: Tushar Mahendra<br>
# date: 2022-09-04</i>
# 
# </p>
# 
# 

# In[2]:


import pandas as pd
import openpyxl
from datetime import date
from datetime import datetime
from os import path
import os


# In[3]:


csv_file = 'C:\\Users\\620110\\OneDrive - Save-On-Foods\\Documents\\Python Project\\Temp_Creator_and_Writer\\Main_Files\\BigData.csv'
master_temp = 'C:\\Users\\620110\\OneDrive - Save-On-Foods\\Documents\\Python Project\\Temp_Creator_and_Writer\\Main_Files\\MasterTemplate.xlsx'


# In[4]:


get_ipython().run_cell_magic('time', '', '\nx = date.today()\n\nprd = "Period 1"\ni = 0\n\nif (x.month == 1):\n    prd = "Period 1"\n    i = 0\n    print("Current Month: January - " + prd)\n\nelif (x.month == 2):\n    prd = "Period 2"\n    i = 1\n    print("Current Month: February - " + prd)\n\nelif (x.month == 3):\n    prd = "Period 3"\n    i = 2\n    print("Current Month: March - " + prd)\n\nelif (x.month == 4):\n    prd = "Period 4"\n    i = 3\n    print("Current Month: April - " + prd)\n\nelif (x.month == 5):\n    prd = "Period 5"\n    i = 4\n    print("Current Month: May - " + prd)\n\nelif (x.month == 6):\n    prd = "Period 6"\n    i = 5\n    print("Current Month: June - " + prd)\n\nelif (x.month == 7):\n    prd = "Period 7"\n    i = 6\n    print("Current Month: July - " + prd)\n\nelif (x.month == 8):\n    prd = "Period 8"\n    i = 7\n    print("Current Month: August - " + prd)\n\nelif (x.month == 9):\n    prd = "Period 9"\n    i = 8\n    print("Current Month: September - " + prd)\n\nelif (x.month == 10):\n    prd = "Period 10"\n    i = 9\n    print("Current Month: October - " + prd)\n    \nelif (x.month == 11):\n    prd = "Period 11"\n    i = 10\n    print("Current Month: November - " + prd)\n\nelif (x.month == 12):\n    prd = "Period 12"\n    i = 11\n    print("Current Month: December - " + prd)')


# In[5]:


#Checks and makes directory for every year
dkr_path = 'C:\\Users\\620110\\OneDrive - Save-On-Foods\\Documents\\Python Project\\Temp_Creator_and_Writer\\Test_Files\\' + str(x.year)
if (path.exists(dkr_path)):
    print("Directory exists.")
    print("Files will be generated in:") 
    print(dkr_path)
    
else:
    (path.os.mkdir(dkr_path)) #Can assign implicit permissions on this directory
    print("Directory does not exist.")
    print("Creating directory: ")
    print(dkr_path)


# In[ ]:


get_ipython().run_cell_magic('time', '', 'df = pd.read_csv(csv_file) #creates a pandas dataframe to read the file\n#print(df)\nnew_df = pd.DataFrame()\n\n\nsplit_values = df[\'User\'].unique() #selects all the unique values from the chosen column.\n#print(split_values)\n# path = \'C:\\\\Users\\\\620110\\\\OneDrive - Save-On-Foods\\\\Documents\\\\Python Project\\\\Mail_Sender\\\\Test_File\\\\file.txt\'\n# split_values.tofile(path, sep=\',\', format=\'\\n%s\')\n\n#Iterates each value from split_values and creates a dataframe of unique column.\nfor value in split_values:\n    store_df = df[df[\'User\'] == value]\n    new_df = store_df[[\'Unpaid Liabilities\', \'Goods In Transit\']]\n    #print(new_df)\n    #print(value)\n    \n    file_path = dkr_path + \'\\\\\' + str(value) + \'.xlsx\'\n    #print(file_path)\n    \n    if(path.exists(file_path)):\n        print("file exists: " + str(value))\n        \n    else:\n        #this creates template for all the stores\n        print("File does not exist. Creating new file: " + str(value))\n        wb_temp = openpyxl.load_workbook(master_temp)\n        wb_temp.save(file_path)\n    \n    #This is where column is being edited   \n    wb = openpyxl.load_workbook(file_path)\n    ws = wb.worksheets[i]\n    ws[\'B2\']=value.replace("SM","")\n    wb.save(file_path)\n        \n    #This is writing the dfs to the excel files\n    writer = pd.ExcelWriter(\n        file_path, \n        if_sheet_exists=\'overlay\', \n        engine=\'openpyxl\', \n        mode=\'a\')\n    \n    new_df.to_excel(writer, sheet_name=prd, startrow=4, startcol=1, index=None, header=False)\n    writer.close()\n    ')


# In[ ]:





# In[30]:


ban_dir = 'C:\\Users\\620110\\OneDrive - Save-On-Foods\\Documents\\Python Project\\Temp_Creator_and_Writer\\Banners\\'

for filename in os.listdir(dkr_path):
    f = os.path.join(dkr_path, filename)
    #checking if it is a file
    if os.path.isfile(f):
#         print(filename)
        edit_filename = filename
        edit_filename = edit_filename.replace('SM.xlsx','')
        zero_filled_number = edit_filename.zfill(4)
        print(zero_filled_number)
#         print(type(zero_filled_number))

        location_id = ("Store " + zero_filled_number)
        print(location_id)
        l = os.path.join(ban_dir,location_id)
        print(l)

        for folder in os.listdir(ban_dir):
            d = os.path.join(ban_dir, folder)
#             print(d)
#             print(folder)

            if(location_id == folder):
                os.replace(f,d)
                print(f + "moved to: " + d)
                
            else:
                print(f + ": new location does not exist.")

 

        
        
    
#     #checking the banner for files
#     if filename <= 1000:
#         print(f)
#     elif filename >= 1001:
#         print(f)

#os.chdir(path)


# In[ ]:


# #This will exit the program and rename the processed file
# y = datetime.now()
# time = y.strftime("%H_%M_%S")
# #print(time)

# proc_file_path = 'C:\\Users\\620110\\OneDrive - Save-On-Foods\\Documents\\Python Project\\Temp_Creator_and_Writer\\Processed_Files\\' + str(x) + '_BigData_' + str(time) +'.csv'
# path.os.rename(csv_file, proc_file_path)
# print(csv_file)
# print("Has moved to:")
# print(proc_file_path)


# In[ ]:





# In[ ]:




